"""Модуль для генерации отчета анализа рисков в Excel."""

from pathlib import Path
from typing import Dict
import shutil
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from logger import get_logger

logger = get_logger("report_generator")


def create_risk_analysis_sheet(
    original_file_path: Path,
    project_params: Dict,
    model_results: Dict,
    risk_analysis: Dict
) -> Path:
    """
    Создание нового листа "Анализ_рисков" в Excel файле.
    
    Args:
        original_file_path: Путь к исходному Excel файлу
        project_params: Параметры проекта
        model_results: Результаты финансовой модели
        risk_analysis: Результаты анализа рисков от ИИ
    
    Returns:
        Путь к сохраненному файлу с анализом
    """
    logger.info(f"Создание листа 'Анализ_рисков' для файла: {original_file_path.name}")
    
    try:
        # Определяем путь для выходного файла
        output_filename = original_file_path.stem + "_анализ.xlsx"
        output_path = original_file_path.parent / output_filename
        
        # Копируем исходный файл, чтобы сохранить все его элементы (формулы, диаграммы, data validation и т.д.)
        logger.debug(f"Копирование исходного файла для сохранения всех элементов")
        shutil.copy2(original_file_path, output_path)
        logger.debug(f"Файл скопирован: {output_path.name}")
        
        # Открываем копию файла в режиме read_only=False для возможности записи
        workbook = load_workbook(output_path, read_only=False)
        
        # Проверка, существует ли уже лист "Анализ_рисков"
        if "Анализ_рисков" in workbook.sheetnames:
            logger.warning("Лист 'Анализ_рисков' уже существует, будет перезаписан")
            workbook.remove(workbook["Анализ_рисков"])
        
        # Создание нового листа
        sheet = workbook.create_sheet("Анализ_рисков")
        
        # Стили для форматирования
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        section_font = Font(bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        wrap_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        row = 1
        
        # Заголовок
        sheet.merge_cells(f'A{row}:D{row}')
        title_cell = sheet[f'A{row}']
        title_cell.value = "Результат анализа рисков проекта"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = center_alignment
        row += 2
        
        # Секция 1: Ключевые показатели
        sheet[f'A{row}'] = "Ключевые показатели проекта"
        sheet[f'A{row}'].font = section_font
        row += 1
        
        # Заголовки таблицы показателей
        headers = ["Параметр", "Значение", "Единица измерения"]
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        row += 1
        
        # Данные показателей
        indicators = [
            ("Тип проекта", project_params.get("type", "Не указан"), ""),
            ("Стоимость строительства", project_params.get("capex", 0), "млн руб"),
            ("Срок строительства", project_params.get("construction_years", 0), "лет"),
            ("Доля долга", project_params.get("debt_share", 0), "%"),
            ("Ставка по долгу", project_params.get("debt_rate", 0), "%"),
            ("Ставка дисконтирования", project_params.get("discount_rate", 0), "%"),
            ("NPV", model_results.get("npv", 0), "млн руб"),
            ("IRR", model_results.get("irr", 0), "%"),
            ("Срок окупаемости", model_results.get("payback_period", 0), "лет")
        ]
        
        for param_name, value, unit in indicators:
            sheet.cell(row=row, column=1).value = param_name
            sheet.cell(row=row, column=1).border = border
            sheet.cell(row=row, column=2).value = value
            sheet.cell(row=row, column=2).border = border
            sheet.cell(row=row, column=3).value = unit
            sheet.cell(row=row, column=3).border = border
            row += 1
        
        row += 2
        
        # Секция 2: Заключение ИИ
        sheet[f'A{row}'] = "Заключение ИИ-анализа"
        sheet[f'A{row}'].font = section_font
        row += 1
        
        # Уровень риска
        risk_level = risk_analysis.get("risk_level", "Не определен")
        sheet[f'A{row}'] = "Уровень риска:"
        sheet[f'A{row}'].font = Font(bold=True)
        sheet[f'B{row}'] = risk_level
        sheet[f'B{row}'].font = Font(bold=True, size=12)
        
        # Цветовая индикация уровня риска
        risk_colors = {
            "Низкий": "00B050",  # Зеленый
            "Средний": "FFC000",  # Желтый
            "Высокий": "FF6600",  # Оранжевый
            "Критический": "C00000"  # Красный
        }
        if risk_level in risk_colors:
            sheet[f'B{row}'].fill = PatternFill(
                start_color=risk_colors[risk_level],
                end_color=risk_colors[risk_level],
                fill_type="solid"
            )
            sheet[f'B{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        
        row += 1
        
        # Обоснование
        sheet[f'A{row}'] = "Обоснование:"
        sheet[f'A{row}'].font = Font(bold=True)
        sheet.merge_cells(f'B{row}:D{row+2}')
        reason_cell = sheet[f'B{row}']
        reason_cell.value = risk_analysis.get("reason", "Не предоставлено")
        reason_cell.alignment = wrap_alignment
        reason_cell.border = border
        row += 3
        
        # Критические факторы
        sheet[f'A{row}'] = "Критические факторы:"
        sheet[f'A{row}'].font = Font(bold=True)
        row += 1
        
        critical_factors = risk_analysis.get("critical_factors", [])
        if critical_factors:
            for factor in critical_factors:
                sheet.cell(row=row, column=1).value = f"• {factor}"
                row += 1
        else:
            sheet.cell(row=row, column=1).value = "Не выявлены"
            row += 1
        
        row += 2
        
        # Секция 3: Анализ сценариев
        scenarios = risk_analysis.get("scenarios", [])
        if scenarios:
            sheet[f'A{row}'] = "Анализ непредвиденных ситуаций"
            sheet[f'A{row}'].font = section_font
            row += 1
            
            # Заголовки таблицы сценариев
            scenario_headers = [
                "Сценарий",
                "Описание",
                "Влияние на NPV, млн руб",
                "Влияние на IRR, %",
                "Вероятность",
                "Потенциальные убытки, млн руб"
            ]
            
            for col_idx, header in enumerate(scenario_headers, start=1):
                cell = sheet.cell(row=row, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            row += 1
            
            # Данные сценариев
            for scenario in scenarios:
                sheet.cell(row=row, column=1).value = scenario.get("name", "Не указан")
                sheet.cell(row=row, column=1).border = border
                
                sheet.cell(row=row, column=2).value = scenario.get("description", "")
                sheet.cell(row=row, column=2).alignment = wrap_alignment
                sheet.cell(row=row, column=2).border = border
                
                npv_impact = scenario.get("npv_impact", 0)
                sheet.cell(row=row, column=3).value = npv_impact
                sheet.cell(row=row, column=3).border = border
                if npv_impact < 0:
                    sheet.cell(row=row, column=3).font = Font(color="C00000")  # Красный для отрицательных
                
                irr_impact = scenario.get("irr_impact", 0)
                sheet.cell(row=row, column=4).value = irr_impact
                sheet.cell(row=row, column=4).border = border
                if irr_impact < 0:
                    sheet.cell(row=row, column=4).font = Font(color="C00000")
                
                sheet.cell(row=row, column=5).value = scenario.get("probability", "Не указана")
                sheet.cell(row=row, column=5).border = border
                
                potential_losses = scenario.get("potential_losses", 0)
                sheet.cell(row=row, column=6).value = potential_losses
                sheet.cell(row=row, column=6).border = border
                if potential_losses > 0:
                    sheet.cell(row=row, column=6).font = Font(color="C00000", bold=True)
                
                row += 1
            
            row += 1
        
        # Секция 4: Суммарные потенциальные убытки
        total_losses = risk_analysis.get("total_potential_losses", 0)
        if total_losses > 0:
            sheet[f'A{row}'] = "Суммарные потенциальные убытки:"
            sheet[f'A{row}'].font = Font(bold=True, size=12)
            sheet[f'B{row}'] = f"{total_losses} млн руб"
            sheet[f'B{row}'].font = Font(bold=True, size=12, color="C00000")
            row += 2
        
        # Секция 5: Рекомендации по снижению рисков
        risk_mitigation = risk_analysis.get("risk_mitigation", [])
        if risk_mitigation:
            sheet[f'A{row}'] = "Рекомендации по снижению рисков:"
            sheet[f'A{row}'].font = section_font
            row += 1
            
            for recommendation in risk_mitigation:
                sheet.cell(row=row, column=1).value = f"• {recommendation}"
                sheet.cell(row=row, column=1).alignment = wrap_alignment
                row += 1
        
        row += 2
        
        # Секция 6: Видение бизнеса
        business_vision = risk_analysis.get("business_vision", "")
        if business_vision:
            sheet[f'A{row}'] = "Предложение дальнейшего видения бизнеса:"
            sheet[f'A{row}'].font = section_font
            row += 1
            
            # Разделяем видение на отдельные предложения для лучшей читаемости
            # Разбиваем по точкам, но сохраняем структуру
            vision_sentences = []
            
            # Сначала пробуем разбить по точкам с пробелом после
            parts = business_vision.split('. ')
            for i, part in enumerate(parts):
                part = part.strip()
                if part:
                    # Если это не последняя часть или последняя часть не заканчивается точкой
                    if i < len(parts) - 1 or not part.endswith('.'):
                        if not part.endswith('.'):
                            vision_sentences.append(part + '.')
                        else:
                            vision_sentences.append(part)
                    else:
                        vision_sentences.append(part)
            
            # Если не удалось разбить на предложения (меньше 2), используем исходный текст
            if len(vision_sentences) <= 1:
                # Пробуем разбить по переносам строк
                vision_sentences = [s.strip() for s in business_vision.split('\n') if s.strip()]
                if len(vision_sentences) <= 1:
                    vision_sentences = [business_vision]
            
            # Выводим каждое предложение с маркером
            for sentence in vision_sentences:
                sentence = sentence.strip()
                if sentence:
                    # Убираем лишние точки в начале
                    if sentence.startswith('.'):
                        sentence = sentence[1:].strip()
                    sheet.cell(row=row, column=1).value = f"• {sentence}"
                    sheet.cell(row=row, column=1).alignment = wrap_alignment
                    # Расширяем ячейку для длинного текста
                    sheet.column_dimensions['A'].width = max(sheet.column_dimensions['A'].width or 0, 80)
                    row += 1
            
            row += 1
        
        # Секция 7: Примерная окупаемость бизнеса
        estimated_payback = risk_analysis.get("estimated_payback")
        if estimated_payback is not None:
            sheet[f'A{row}'] = "Примерная окупаемость бизнеса (с учетом перспектив развития):"
            sheet[f'A{row}'].font = Font(bold=True)
            sheet[f'B{row}'] = f"{estimated_payback} лет"
            sheet[f'B{row}'].font = Font(bold=True, size=12)
            row += 2
        
        # Автоподбор ширины столбцов
        for col in range(1, 7):
            max_length = 0
            column_letter = get_column_letter(col)
            for cell in sheet[column_letter]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Сохранение файла
        try:
            workbook.save(output_path)
            logger.debug("Файл успешно сохранен")
        except Exception as save_error:
            workbook.close()
            logger.error(f"Ошибка при сохранении файла: {save_error}", exc_info=True)
            # Удаляем поврежденную копию при ошибке
            if output_path.exists():
                try:
                    output_path.unlink()
                except:
                    pass
            raise ValueError(f"Ошибка сохранения файла: {str(save_error)}")
        
        workbook.close()
        
        # Проверяем, что файл создан и не пустой
        if not output_path.exists():
            raise ValueError("Файл не был создан после сохранения")
        
        file_size = output_path.stat().st_size
        if file_size == 0:
            output_path.unlink()
            raise ValueError("Созданный файл пустой")
        
        if file_size < 1000:  # Минимальный размер для валидного Excel файла
            logger.warning(f"Файл очень маленький ({file_size} байт), возможно поврежден")
        
        # Пробуем открыть файл для проверки валидности
        try:
            test_workbook = load_workbook(output_path, read_only=True, data_only=True)
            test_workbook.close()
            logger.debug("Файл успешно прошел валидацию")
        except Exception as validation_error:
            logger.error(f"Файл не прошел валидацию: {validation_error}")
            # Удаляем поврежденный файл
            if output_path.exists():
                output_path.unlink()
            raise ValueError(f"Созданный файл поврежден и не может быть открыт: {validation_error}")
        
        logger.info(f"Отчет успешно создан: {output_path.name} (размер: {file_size} байт)")
        return output_path
        
    except Exception as e:
        logger.error(f"Ошибка при создании отчета: {e}", exc_info=True)
        raise ValueError(f"Ошибка генерации отчета: {str(e)}")

