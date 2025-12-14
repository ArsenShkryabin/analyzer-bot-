"""Модуль для извлечения данных из Excel файлов."""

from pathlib import Path
from typing import Dict, Optional, Tuple
import openpyxl
from openpyxl import load_workbook

from logger import get_logger

logger = get_logger("excel_reader")


def find_cell_by_text(worksheet, search_text: str, case_sensitive: bool = False) -> Optional[Tuple[int, int]]:
    """
    Поиск ячейки по тексту в первом листе.
    
    Args:
        worksheet: Рабочий лист openpyxl
        search_text: Текст для поиска
        case_sensitive: Учитывать регистр при поиске
    
    Returns:
        Кортеж (row, column) или None, если не найдено
    """
    search_lower = search_text.lower() if not case_sensitive else search_text
    
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value:
                cell_value = str(cell.value).lower() if not case_sensitive else str(cell.value)
                if search_lower in cell_value:
                    logger.debug(f"Найдена ячейка с текстом '{search_text}': {cell.coordinate}")
                    return (cell.row, cell.column)
    
    return None


def extract_value_near_cell(worksheet, row: int, col: int, offset: int = 1) -> Optional[float]:
    """
    Извлечение числового значения из ячейки рядом с найденной.
    
    Args:
        worksheet: Рабочий лист openpyxl
        row: Номер строки найденной ячейки
        col: Номер столбца найденной ячейки
        offset: Смещение для поиска значения (по умолчанию следующая ячейка справа)
    
    Returns:
        Числовое значение или None
    """
    # Пробуем разные позиции: справа, справа+1, справа+2, снизу, снизу+1, снизу+2
    search_positions = [
        (row, col + 1),      # Справа
        (row, col + 2),      # Справа через одну
        (row, col + 3),      # Справа через две
        (row + 1, col),      # Снизу
        (row + 1, col + 1),  # Снизу справа
        (row + 1, col + 2),  # Снизу справа через одну
        (row + 2, col),      # Снизу через одну
        (row + 2, col + 1),  # Снизу через одну справа
    ]
    
    for r, c in search_positions:
        try:
            value_cell = worksheet.cell(row=r, column=c)
            if value_cell.value is not None:
                # Пробуем преобразовать в число
                try:
                    value = float(value_cell.value)
                    logger.debug(f"Найдено значение {value} в ячейке {value_cell.coordinate}")
                    return value
                except (ValueError, TypeError):
                    # Если не число, пробуем извлечь число из строки
                    if isinstance(value_cell.value, str):
                        import re
                        # Ищем числа в строке (включая проценты)
                        numbers = re.findall(r'-?\d+\.?\d*', value_cell.value.replace(',', '.').replace('%', ''))
                        if numbers:
                            try:
                                value = float(numbers[0])
                                logger.debug(f"Извлечено значение {value} из строки '{value_cell.value}'")
                                return value
                            except ValueError:
                                pass
        except Exception:
            continue
    
    # Если не нашли рядом, ищем в той же строке или столбце в пределах 10 ячеек
    for i in range(1, 11):
        try:
            # Ищем справа
            value_cell = worksheet.cell(row=row, column=col + i)
            if value_cell.value is not None:
                try:
                    value = float(value_cell.value)
                    logger.debug(f"Найдено значение {value} в ячейке {value_cell.coordinate} (поиск в строке)")
                    return value
                except (ValueError, TypeError):
                    pass
        except:
            pass
        
        try:
            # Ищем снизу
            value_cell = worksheet.cell(row=row + i, column=col)
            if value_cell.value is not None:
                try:
                    value = float(value_cell.value)
                    logger.debug(f"Найдено значение {value} в ячейке {value_cell.coordinate} (поиск в столбце)")
                    return value
                except (ValueError, TypeError):
                    pass
        except:
            pass
    
    return None


def extract_project_data(file_path: Path) -> Dict:
    """
    Извлечение данных проекта из Excel файла.
    
    Args:
        file_path: Путь к Excel файлу
    
    Returns:
        Словарь с извлеченными данными
    
    Raises:
        ValueError: Если файл не удалось прочитать или не найдены необходимые данные
    """
    logger.info(f"Начало извлечения данных из файла: {file_path.name}")
    
    try:
        workbook = load_workbook(file_path, data_only=True)
        
        # Получаем список всех листов
        sheet_names = workbook.sheetnames
        logger.debug(f"Найдено листов в файле: {len(sheet_names)}: {', '.join(sheet_names)}")
        
        data = {}
        
        # Ищем данные на всех листах
        for sheet_name in sheet_names:
            worksheet = workbook[sheet_name]
            logger.debug(f"Обработка листа: {worksheet.title}")
        
            # Поиск параметров проекта (только на первом листе)
            if sheet_name == sheet_names[0]:
                # Тип проекта - ищем "тип проекта" или просто значение "Метро"
                project_type_cell = find_cell_by_text(worksheet, "тип проекта", case_sensitive=False)
                if not project_type_cell:
                    # Пробуем найти просто "Метро" или другие типы
                    for project_type in ["метро", "дорога", "энергетика", "мост", "тоннель"]:
                        project_type_cell = find_cell_by_text(worksheet, project_type, case_sensitive=False)
                        if project_type_cell:
                            row, col = project_type_cell
                            cell = worksheet.cell(row=row, column=col)
                            if cell.value and isinstance(cell.value, str):
                                data["type"] = cell.value.strip()
                                logger.debug(f"Тип проекта: {data['type']} (лист: {sheet_name})")
                                break
                
                if project_type_cell and "type" not in data:
                    # Пробуем извлечь тип проекта из соседних ячеек
                    row, col = project_type_cell
                    for offset in [1, 2, -1, -2]:
                        try:
                            cell = worksheet.cell(row=row, column=col + offset)
                            if cell.value and isinstance(cell.value, str) and len(cell.value) > 2:
                                # Пропускаем служебные тексты
                                if cell.value.strip().upper() not in ["ПАРАМЕТРЫ ПРОЕКТА", "ПРОЕКТ", "ТИП"]:
                                    data["type"] = cell.value.strip()
                                    logger.debug(f"Тип проекта: {data['type']} (лист: {sheet_name})")
                                    break
                        except:
                            continue
                
                if "type" not in data:
                    data["type"] = "Не указан"
            
            # Поиск числовых параметров на всех листах
            search_patterns = {
                "capex": ["стоимость строительства", "capex", "капитальные затраты", "инвестиции"],
                "construction_years": ["срок строительства", "период строительства", "длительность строительства"],
                "debt_share": ["доля долга", "процент долга", "доля заемных средств"],
                "debt_rate": ["ставка по долгу", "ставка долга", "процент по долгу", "ставка займа", "процент долга"],
                "discount_rate": ["ставка дисконтирования", "дисконт", "ставка дисконта", "ставка дисконтирования для бизнеса"],
                "npv": ["npv", "чистая приведенная стоимость", "чистая приведённая стоимость", "чистый приведенный доход", "чистая приведённая стоимость, млн. руб", "чистая приведённая стоимость, млн руб"],
                "irr": ["irr", "внутренняя норма доходности", "внутренняя норма рентабельности"],
                "payback_period": ["срок окупаемости", "период окупаемости", "окупаемость", "payback", "окупаемость, лет"]
            }
            
            # Ищем параметры, которые еще не найдены
            for key, patterns in search_patterns.items():
                # Пропускаем, если параметр уже найден
                if key in data and data[key] != 0.0:
                    continue
                    
                found = False
                
                # Специальная обработка для IRR - проверяем E15 (значение рядом с D15, где текст)
                if key == "irr" and sheet_name == sheet_names[0]:
                    try:
                        # Проверяем E15 (столбец 5, строка 15) - значение рядом с текстом в D15
                        cell = worksheet.cell(row=15, column=5)  # E = 5
                        if cell.value is not None:
                            try:
                                value = float(cell.value)
                                # Если значение меньше 1, это десятичная дробь (0.30 = 30%), умножаем на 100
                                if 0 <= value < 1:
                                    value = value * 100
                                if 0 <= value <= 100:  # IRR обычно в процентах
                                    data[key] = value
                                    logger.debug(f"{key}: {value} (лист: {sheet_name}, ячейка E15)")
                                    found = True
                            except (ValueError, TypeError):
                                # Пробуем извлечь из строки (например, "30%")
                                if isinstance(cell.value, str):
                                    import re
                                    numbers = re.findall(r'\d+\.?\d*', cell.value.replace(',', '.').replace('%', ''))
                                    if numbers:
                                        val = float(numbers[0])
                                        # Если значение меньше 1, умножаем на 100
                                        if val < 1:
                                            val = val * 100
                                        if 0 <= val <= 100:
                                            data[key] = val
                                            logger.debug(f"{key}: {val} (лист: {sheet_name}, ячейка E15, из текста '{cell.value}')")
                                            found = True
                        
                        # Также проверяем D15 и D16 на случай, если значение там
                        if not found:
                            for row_num in [15, 16]:
                                cell = worksheet.cell(row=row_num, column=4)  # D = 4
                                if cell.value is not None:
                                    try:
                                        value = float(cell.value)
                                        if 0 <= value <= 100:
                                            data[key] = value
                                            logger.debug(f"{key}: {value} (лист: {sheet_name}, ячейка D{row_num})")
                                            found = True
                                            break
                                    except (ValueError, TypeError):
                                        if isinstance(cell.value, str):
                                            import re
                                            numbers = re.findall(r'\d+\.?\d*', cell.value.replace(',', '.').replace('%', ''))
                                            if numbers:
                                                val = float(numbers[0])
                                                if 0 <= val <= 100:
                                                    data[key] = val
                                                    logger.debug(f"{key}: {val} (лист: {sheet_name}, ячейка D{row_num}, из текста)")
                                                    found = True
                                                    break
                    except Exception as e:
                        logger.debug(f"Ошибка при проверке D15-E15 для IRR: {e}")
                
                if found:
                    continue
                    
                for pattern in patterns:
                    cell_pos = find_cell_by_text(worksheet, pattern, case_sensitive=False)
                    if cell_pos:
                        # Сначала проверяем саму найденную ячейку на наличие числа
                        row, col = cell_pos
                        cell = worksheet.cell(row=row, column=col)
                        if cell.value:
                            # Пробуем извлечь число из самой ячейки
                            if isinstance(cell.value, (int, float)):
                                data[key] = float(cell.value)
                                logger.debug(f"{key}: {data[key]} (лист: {sheet_name}, из самой ячейки)")
                                found = True
                                break
                            elif isinstance(cell.value, str):
                                import re
                                # Ищем числа в строке (включая проценты)
                                numbers = re.findall(r'-?\d+\.?\d*', cell.value.replace(',', '.').replace('%', ''))
                                if numbers:
                                    try:
                                        value = float(numbers[0])
                                        data[key] = value
                                        logger.debug(f"{key}: {value} (лист: {sheet_name}, из текста ячейки)")
                                        found = True
                                        break
                                    except ValueError:
                                        pass
                        
                        # Если не нашли в самой ячейке, ищем рядом
                        value = extract_value_near_cell(worksheet, row, col)
                        if value is not None:
                            data[key] = value
                            logger.debug(f"{key}: {value} (лист: {sheet_name})")
                            found = True
                            break
                
                if not found and key not in data:
                    data[key] = 0.0
        
        workbook.close()
        
        # Логируем предупреждения для не найденных параметров
        for key in search_patterns.keys():
            if key not in data or data[key] == 0.0:
                logger.warning(f"Не удалось найти параметр: {key}")
        
        # Проверка наличия критических данных
        required_fields = ["npv", "irr", "payback_period"]
        missing_fields = [field for field in required_fields if field not in data or data[field] == 0.0]
        
        if missing_fields:
            raise ValueError(
                f"В файле не найдены необходимые данные: {', '.join(missing_fields)}. "
                "Проверьте структуру файла."
            )
        
        logger.info("Данные успешно извлечены из файла")
        return data
        
    except openpyxl.utils.exceptions.InvalidFileException:
        raise ValueError("Не удалось прочитать файл. Убедитесь, что файл не поврежден.")
    except Exception as e:
        logger.error(f"Ошибка при извлечении данных: {e}", exc_info=True)
        raise ValueError(f"Ошибка обработки файла: {str(e)}")

